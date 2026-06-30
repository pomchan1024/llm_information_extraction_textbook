! pip install openpyxl==3.1.5

import openpyxl

def extract_text_from_excel(file_path):
    # ファイルを読み込む（data_only=Trueで、計算式ではなく計算結果を取得）
    wb = openpyxl.load_workbook(file_path, data_only=True)
    full_text = []

    # すべてのシートをループ
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        full_text.append(f"--- Sheet: {sheet_name} ---")  # シート名を区切りとして追加

        # すべての行と列をスキャン
        for row in sheet.iter_rows(values_only=True):
            # None（空セル）を除外して文字列に変換
            row_text = [str(cell) for cell in row if cell is not None]
            if row_text:
                full_text.append(" ".join(row_text))

    return "\n".join(full_text)


# --- 実行例 ---
text = extract_text_from_excel("sample_data/sample.xlsx")
print(text)